import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from datetime import datetime


def process_workbook(file):
    workbook = xl.load_workbook(file)
    sheet = workbook['Sheet1']

    sheet.cell(1, 6).value = 'Discount'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        unit = sheet.cell(row, 3)
        amount = sheet.cell(row, 4)
        total = sheet.cell(row, 5)
        total.value = unit.value * amount.value
        
        discount = total.value * 0.1 
        discount_cell = sheet.cell(row, 6)
        discount_cell.value = discount
        

    values = Reference(sheet, min_row=2, 
            max_row=sheet.max_row,
            min_col=5, 
            max_col=5)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'g2')


    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'transactions_{timestamp}.xlsx'
    workbook.save(filename)
    return filename


file_name = process_workbook('transactions.xlsx')

print(f'success! Saved as {file_name}')